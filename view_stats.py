import os
import sys
import argparse
import subprocess
from openpyxl import load_workbook

# -------------------
# CLI arguments
# -------------------
parser = argparse.ArgumentParser(description="View stats deltas for a player")
parser.add_argument("-ign", "--username", required=True, help="Minecraft IGN")
args = parser.parse_args()

USERNAME = args.username
EXCEL_FILE = "sheep_wars_stats.xlsx"

# -------------------
# Fetch latest stats via get.py (no lifetime logging)
# -------------------
subprocess.run([sys.executable, "get.py", "-ign", USERNAME, "-nolifetime"], capture_output=True, text=True)

# -------------------
# Load workbook
# -------------------
if not os.path.exists(EXCEL_FILE):
    raise RuntimeError("Excel file not found")

wb = load_workbook(EXCEL_FILE)

if USERNAME not in wb.sheetnames:
    # try case-insensitive match
    key = USERNAME.casefold()
    found = None
    for s in wb.sheetnames:
        if s.casefold() == key:
            found = s
            break
    if found:
        player_ws = wb[found]
    else:
        raise RuntimeError("Player sheet not found")
else:
    player_ws = wb[USERNAME]

# helper to read snapshot values from D/E starting at given row
def read_snapshot(start_row):
    stat_names = ["Kills", "Deaths", "K/D", "Wins", "Losses", "W/L"]
    out = {}
    for i, name in enumerate(stat_names):
        val = player_ws[f"E{start_row + i}"].value
        if val is None:
            return None
        out[name] = val
    return out

# read current all-time values from B (rows 39-44)
stat_names = ["Kills", "Deaths", "K/D", "Wins", "Losses", "W/L"]
all_time_start_row = 39
all_time = {}
for i, name in enumerate(stat_names):
    all_time[name] = player_ws[f"B{all_time_start_row + i}"].value or 0

# periods: (snapshot_start_row, target_start_row)
periods = {
    "Session": (3, 3),
    "Daily": (12, 12),
    "Weekly": (21, 21),
    "Monthly": (30, 30),
}

for period, (snap_row, target_row) in periods.items():
    snap = read_snapshot(snap_row)
    if snap is None:
        # no snapshot for this period; skip updating deltas
        continue

    # compute deltas and write into column B at target rows
    # extract numeric deltas for counts
    try:
        kills_delta = (all_time.get("Kills", 0) or 0) - (snap.get("Kills", 0) or 0)
        deaths_delta = (all_time.get("Deaths", 0) or 0) - (snap.get("Deaths", 0) or 0)
        wins_delta = (all_time.get("Wins", 0) or 0) - (snap.get("Wins", 0) or 0)
        losses_delta = (all_time.get("Losses", 0) or 0) - (snap.get("Losses", 0) or 0)
    except Exception:
        kills_delta = deaths_delta = wins_delta = losses_delta = 0

    # write counts
    player_ws[f"B{target_row + 0}"] = kills_delta
    player_ws[f"B{target_row + 1}"] = deaths_delta

    # K/D ratio for the period (use deltas)
    if deaths_delta and deaths_delta != 0:
        kd_ratio = kills_delta / deaths_delta
    else:
        kd_ratio = float(kills_delta) if kills_delta else 0.0
    player_ws[f"B{target_row + 2}"] = round(kd_ratio, 2)

    # write wins/losses and W/L ratio
    player_ws[f"B{target_row + 3}"] = wins_delta
    player_ws[f"B{target_row + 4}"] = losses_delta
    if losses_delta and losses_delta != 0:
        wl_ratio = wins_delta / losses_delta
    else:
        wl_ratio = float(wins_delta) if wins_delta else 0.0
    player_ws[f"B{target_row + 5}"] = round(wl_ratio, 2)

# Write all-time current values into B (so column B shows the current cumulative values)
for i, name in enumerate(stat_names):
    player_ws[f"B{all_time_start_row + i}"] = all_time.get(name, 0)

wb.save(EXCEL_FILE)
