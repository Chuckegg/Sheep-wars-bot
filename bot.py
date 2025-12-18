import discord
from discord.ext import commands
import subprocess
import sys
from openpyxl import load_workbook
import os
import re
from zoneinfo import ZoneInfo
import json

# sanitize output for Discord (remove problematic unicode/control chars)
def sanitize_output(text: str) -> str:
    if text is None:
        return ""
    # Replace a few common emoji with ASCII labels
    replacements = {
        '✅': '[OK]',
        '❌': '[ERROR]',
        '⚠️': '[WARNING]',
        '📊': '[DATA]',
        '📋': '[INFO]',
        '⏭️': '[SKIP]',
    }
    for k, v in replacements.items():
        text = text.replace(k, v)

    # Remove C0 control chars except newline and tab
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', str(text))
    # Collapse very long whitespace
    text = re.sub(r"\s{3,}", ' ', text)
    return text

# additional imports for background tasks
import asyncio
import datetime

# tracked users file and creator identifier
TRACKED_FILE = os.path.join(os.path.dirname(__file__), "tracked_users.txt")
USER_LINKS_FILE = os.path.join(os.path.dirname(__file__), "user_links.json")
CREATOR_NAME = "chuckegg"  # case-insensitive match fallback
# Optionally set a numeric Discord user ID for direct DM (recommended for reliability)
# Example: CREATOR_ID = 123456789012345678
CREATOR_ID = "542467909549555734"
CREATOR_TZ = ZoneInfo("America/New_York")

# Prestige icons per 100 levels (index 0 = levels 0-99)
PRESTIGE_ICONS = [
    "❤", "✙", "✫", "✈", "✠", "♙", "⚡", "☢", "✏", "☯",
    "☃️", "۞", "✤", "♫", "♚", "❉", "Σ", "￡", "✖", "❁",
    "✚", "✯", "✆", "❥", "☾⋆⁺", "⚜", "✦", "⚝", "✉", "ツ",
    "❣", "✮", "✿", "✲", "❂", "ƒ", "$", "⋚⋚", "Φ", "✌",
]

# Prestige colors (RGB tuples for Discord embed colors)
# Levels: 0, 100, 200, 300, 400, 500, 600, 700, 800, 900, 1000+
PRESTIGE_COLORS = {
    0: (119, 119, 119),      # GRAY (§7)
    100: (255, 255, 255),    # WHITE (§f)
    200: (255, 85, 85),      # RED (§c)
    300: (255, 170, 0),      # GOLD (§6)
    400: (255, 255, 85),     # YELLOW (§e)
    500: (85, 255, 85),      # LIGHT_GREEN (§a)
    600: (0, 170, 170),      # DARK_AQUA (§3)
    700: (170, 0, 170),      # DARK_PURPLE (§5)
    800: (255, 85, 255),     # LIGHT_PURPLE (§d)
    900: None,               # Rainbow (special handling)
    1000: (255, 255, 255),   # WHITE (§f)
    1100: (255, 255, 255),   # WHITE brackets and numbers
    1200: (255, 85, 85),     # RED brackets and numbers
    1300: (255, 170, 0),     # GOLD/ORANGE brackets and numbers
    1400: (255, 255, 85),    # YELLOW brackets and numbers
    1500: (85, 255, 85),     # GREEN brackets and numbers
    1600: (85, 255, 255),    # CYAN brackets and numbers
    1700: (255, 85, 255),    # MAGENTA brackets and numbers
    1800: (255, 85, 255),    # PINK/MAGENTA brackets and numbers
    1900: None,              # Rainbow (special handling)
    2000: (170, 170, 170),   # GRAY/TAN brackets and numbers
    2100: (255, 255, 255),   # WHITE brackets with gray numbers
    2200: (255, 85, 85),     # RED brackets with yellow numbers
    2300: None,              # Rainbow brackets
    2400: (170, 0, 170),     # PURPLE brackets with green numbers
    2500: (255, 255, 255),   # WHITE brackets with green numbers
    2600: (255, 255, 255),   # WHITE brackets with cyan numbers
    2700: (255, 255, 255),   # WHITE brackets with magenta numbers
    2800: (255, 85, 85),     # RED brackets with dark red numbers
    2900: None,              # Rainbow brackets
    3000: (255, 255, 255),   # WHITE brackets with gray numbers
    3100: (255, 255, 255),   # WHITE brackets and numbers
    3200: (255, 85, 85),     # RED brackets and numbers
    3300: None,              # Rainbow brackets (orange/red/yellow)
    3400: None,              # Rainbow brackets (yellow/orange)
    3500: (85, 255, 85),     # GREEN brackets and numbers
    3600: (85, 255, 255),    # CYAN/BLUE brackets and numbers
    3700: (255, 255, 255),   # WHITE/YELLOW brackets with magenta numbers
    3800: None,              # Rainbow brackets (purple/red)
    3900: None,              # Rainbow brackets (full spectrum)
    4000: (255, 255, 255),   # WHITE brackets with black numbers
}


def get_prestige_icon(level: int) -> str:
    try:
        lvl = int(level)
    except Exception:
        lvl = 0
    idx = max(0, lvl // 100)
    if idx >= len(PRESTIGE_ICONS):
        idx = len(PRESTIGE_ICONS) - 1
    return PRESTIGE_ICONS[idx]

def get_prestige_color(level: int) -> tuple:
    """Get RGB color tuple for a given prestige level.
    Supports levels 0-1000. Returns default dark gray for levels outside this range.
    """
    try:
        lvl = int(level)
    except Exception:
        lvl = 0
    
    # Find the closest prestige level color
    for prestige_level in sorted(PRESTIGE_COLORS.keys(), reverse=True):
        if lvl >= prestige_level:
            color = PRESTIGE_COLORS[prestige_level]
            # Handle Rainbow (None) by returning a default color or cycling
            if color is None:
                # For now, return a vibrant color for rainbow
                return (255, 100, 200)
            return color
    
    # Fallback to gray if below 0
    return (119, 119, 119)

def get_ansi_color_code(level: int) -> str:
    """Get ANSI color code for a given prestige level."""
    color = get_prestige_color(level)
    
    # Map RGB to closest basic ANSI color for Discord compatibility
    r, g, b = color
    
    # Determine which basic ANSI color is closest
    if r > 200 and g > 200 and b > 200:
        return "\u001b[0;37m"  # White
    elif r < 100 and g < 100 and b < 100:
        return "\u001b[0;30m"  # Gray
    elif r > 200 and g < 100 and b < 100:
        return "\u001b[0;31m"  # Red
    elif r > 200 and g > 150 and b < 100:
        return "\u001b[0;33m"  # Yellow/Gold
    elif r < 100 and g > 200 and b < 100:
        return "\u001b[0;32m"  # Green
    elif r < 100 and g > 150 and b > 150:
        return "\u001b[0;36m"  # Cyan
    elif r > 150 and g < 100 and b > 150:
        return "\u001b[0;35m"  # Magenta/Pink
    elif r > 200 and g > 200 and b < 100:
        return "\u001b[0;33m"  # Yellow
    else:
        return "\u001b[0;37m"  # Default White

def load_tracked_users():
    if not os.path.exists(TRACKED_FILE):
        return []
    with open(TRACKED_FILE, "r", encoding="utf-8") as f:
        lines = [l.strip() for l in f.readlines() if l.strip()]
    return lines

def add_tracked_user(ign: str) -> bool:
    users = load_tracked_users()
    key = ign.casefold()
    for u in users:
        if u.casefold() == key:
            return False
    # append
    with open(TRACKED_FILE, "a", encoding="utf-8") as f:
        f.write(ign + "\n")
    return True

def load_user_links():
    """Load username -> Discord user ID mappings from JSON file"""
    if not os.path.exists(USER_LINKS_FILE):
        return {}
    try:
        with open(USER_LINKS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_user_links(links: dict):
    """Save username -> Discord user ID mappings to JSON file"""
    with open(USER_LINKS_FILE, "w", encoding="utf-8") as f:
        json.dump(links, f, indent=2)

def link_user_to_ign(discord_user_id: int, ign: str):
    """Link a Discord user ID to a Minecraft username (case-insensitive)"""
    links = load_user_links()
    # Store with original case but search case-insensitively
    links[ign.casefold()] = str(discord_user_id)
    save_user_links(links)

def is_user_authorized(discord_user_id: int, ign: str) -> bool:
    """Check if a Discord user is authorized to manage a username"""
    links = load_user_links()
    key = ign.casefold()
    return links.get(key) == str(discord_user_id)

def remove_tracked_user(ign: str) -> bool:
    """Remove a username from tracked users list"""
    users = load_tracked_users()
    key = ign.casefold()
    found = False
    new_users = []
    for u in users:
        if u.casefold() == key:
            found = True
        else:
            new_users.append(u)
    
    if found:
        with open(TRACKED_FILE, "w", encoding="utf-8") as f:
            for u in new_users:
                f.write(u + "\n")
    return found

def unlink_user_from_ign(ign: str) -> bool:
    """Remove username -> Discord user ID link"""
    links = load_user_links()
    key = ign.casefold()
    if key in links:
        del links[key]
        save_user_links(links)
        return True
    return False

async def run_get_for_users(flag: str):
    users = load_tracked_users()
    if not users:
        return users
    fetched = []
    for u in users:
        try:
            # run synchronously in thread to avoid blocking loop
            def call_combined():
                # Single call: snapshot flag + refresh
                return subprocess.run([sys.executable, "get.py", flag, "-refresh", "-ign", u], capture_output=True, text=True)
            await asyncio.to_thread(call_combined)
            fetched.append(u)
        except Exception:
            continue
    return fetched

async def run_get_for_users_multi(flags: list[str]):
    users = load_tracked_users()
    if not users:
        return users
    fetched = []
    for u in users:
        try:
            def call_multi():
                # Single call per user: all flags + refresh
                return subprocess.run([sys.executable, "get.py", *flags, "-refresh", "-ign", u], capture_output=True, text=True)
            await asyncio.to_thread(call_multi)
            fetched.append(u)
        except Exception:
            continue
    return fetched

async def send_fetch_message(message: str):
    # DM the creator (prefer explicit ID if set)
    user = None
    if CREATOR_ID is not None:
        try:
            uid = int(CREATOR_ID)
            user = bot.get_user(uid) or await bot.fetch_user(uid)
        except Exception:
            user = None
    if user is None:
        # fallback to name/display name search across guilds
        for guild in bot.guilds:
            for member in guild.members:
                if member.bot:
                    continue
                name_match = member.name.casefold() == CREATOR_NAME.casefold()
                display_match = member.display_name.casefold() == CREATOR_NAME.casefold()
                if name_match or display_match:
                    user = member
                    break
            if user:
                break
    if user:
        try:
            await user.send(message)
            return
        except Exception as e:
            # Common cause: user has DMs disabled (Discord error 50007). Fall back to channel.
            print(f"[WARNING] Could not DM creator: {e}")
    # fallback: send to system channel or first writable channel
    for guild in bot.guilds:
        channel = None
        if guild.system_channel and guild.system_channel.permissions_for(guild.me).send_messages:
            channel = guild.system_channel
        else:
            for ch in guild.text_channels:
                if ch.permissions_for(guild.me).send_messages:
                    channel = ch
                    break
        if channel:
            try:
                await channel.send(message)
                break
            except Exception:
                continue

async def scheduler_loop():
    last_daily = None
    last_weekly = None
    last_monthly = None
    while True:
        now = datetime.datetime.now(tz=CREATOR_TZ)
        # run daily at 9:30
        if now.hour == 9 and now.minute == 30:
            today = now.date()
            if last_daily != today:
                fetched = await run_get_for_users("-daily")
                if fetched:
                    await send_fetch_message(f"Fetched -daily for usernames {', '.join(fetched)}.")
                last_daily = today
            # weekly check (Monday)
            if now.weekday() == 0:
                iso_week = now.isocalendar()[1]
                if last_weekly != iso_week:
                    fetched = await run_get_for_users("-weekly")
                    if fetched:
                        await send_fetch_message(f"Fetched -weekly for usernames {', '.join(fetched)}.")
                    last_weekly = iso_week
            # monthly check (day 1)
            if now.day == 1:
                month = (now.year, now.month)
                if last_monthly != month:
                    fetched = await run_get_for_users("-monthly")
                    if fetched:
                        await send_fetch_message(f"Fetched -monthly for usernames {', '.join(fetched)}.")
                    last_monthly = month

        await asyncio.sleep(20)

# Helper class for stats tab view
class StatsTabView(discord.ui.View):
    def __init__(self, sheet, ign, level_value: int, prestige_icon: str):
        super().__init__()
        self.sheet = sheet
        self.ign = ign
        self.level_value = level_value
        self.prestige_icon = prestige_icon
        self.current_tab = "all-time"
        
        # Row mappings: (kills_row, deaths_row, kd_row, wins_row, losses_row, wl_row)
        self.tabs = {
            "all-time": (39, 40, 41, 42, 43, 44),
            "session": (3, 4, 5, 6, 7, 8),
            "daily": (12, 13, 14, 15, 16, 17),
            "weekly": (21, 22, 23, 24, 25, 26),
            "monthly": (30, 31, 32, 33, 34, 35),
        }
        self.update_buttons()
    
    def update_buttons(self):
        # Update button styles based on current tab
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                if child.custom_id == self.current_tab:
                    child.style = discord.ButtonStyle.primary
                else:
                    child.style = discord.ButtonStyle.secondary
    
    def get_stats_embed(self, tab_name):
        rows = self.tabs[tab_name]
        kills = self.sheet[f"B{rows[0]}"].value or 0
        deaths = self.sheet[f"B{rows[1]}"].value or 0
        kd_ratio = self.sheet[f"B{rows[2]}"].value or 0
        wins = self.sheet[f"B{rows[3]}"].value or 0
        losses = self.sheet[f"B{rows[4]}"].value or 0
        wl_ratio = self.sheet[f"B{rows[5]}"].value or 0
        
        # Get prestige color based on level
        prestige_color = get_prestige_color(self.level_value)
        ansi_code = get_ansi_color_code(self.level_value)
        reset_code = "\u001b[0;0m"
        
        embed = discord.Embed(
            title="",
            color=discord.Color.from_rgb(*prestige_color)
        )
        
        # Add colored level display with full title as a full-width field
        # Only the level and icon inside brackets are colored
        colored_title = f"[{ansi_code}{self.level_value}{self.prestige_icon}{reset_code}] {self.ign} - {tab_name.title()} Stats"
        embed.add_field(name="", value=f"```ansi\n{colored_title}```", inline=False)
        
        # Add 6 inline fields: label as field name, data in compact code block
        embed.add_field(name="Wins", value=f"```{str(wins)}```", inline=True)
        embed.add_field(name="Losses", value=f"```{str(losses)}```", inline=True)
        embed.add_field(name="W/L Ratio", value=f"```{str(wl_ratio)}```", inline=True)

        embed.add_field(name="Kills", value=f"```{str(kills)}```", inline=True)
        embed.add_field(name="Deaths", value=f"```{str(deaths)}```", inline=True)
        embed.add_field(name="K/D Ratio", value=f"```{str(kd_ratio)}```", inline=True)
        
        return embed
    
    @discord.ui.button(label="All-time", custom_id="all-time", style=discord.ButtonStyle.primary)
    async def all_time_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "all-time"
        self.update_buttons()
        embed = self.get_stats_embed(self.current_tab)
        await interaction.response.edit_message(embed=embed, view=self)
    
    @discord.ui.button(label="Session", custom_id="session", style=discord.ButtonStyle.secondary)
    async def session_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "session"
        self.update_buttons()
        embed = self.get_stats_embed(self.current_tab)
        await interaction.response.edit_message(embed=embed, view=self)
    
    @discord.ui.button(label="Daily", custom_id="daily", style=discord.ButtonStyle.secondary)
    async def daily_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "daily"
        self.update_buttons()
        embed = self.get_stats_embed(self.current_tab)
        await interaction.response.edit_message(embed=embed, view=self)
    
    @discord.ui.button(label="Weekly", custom_id="weekly", style=discord.ButtonStyle.secondary)
    async def weekly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "weekly"
        self.update_buttons()
        embed = self.get_stats_embed(self.current_tab)
        await interaction.response.edit_message(embed=embed, view=self)
    
    @discord.ui.button(label="Monthly", custom_id="monthly", style=discord.ButtonStyle.secondary)
    async def monthly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "monthly"
        self.update_buttons()
        embed = self.get_stats_embed(self.current_tab)
        await interaction.response.edit_message(embed=embed, view=self)


# Leaderboard view for switching between periods
class LeaderboardView(discord.ui.View):
    def __init__(self, metric: str, wb):
        super().__init__()
        self.metric = metric  # "kills", "deaths", "kdr", "wins", "losses", "wlr"
        self.wb = wb
        self.current_period = "lifetime"
        
        # Row mappings for column B: (kills, deaths, kdr, wins, losses, wlr)
        self.periods = {
            "lifetime": (39, 40, 41, 42, 43, 44),
            "session": (3, 4, 5, 6, 7, 8),
            "daily": (12, 13, 14, 15, 16, 17),
            "weekly": (21, 22, 23, 24, 25, 26),
            "monthly": (30, 31, 32, 33, 34, 35),
        }
        # Map metric names to index in rows tuple
        self.metric_indices = {
            "kills": 0,
            "deaths": 1,
            "kdr": 2,
            "wins": 3,
            "losses": 4,
            "wlr": 5,
        }
        self.metric_labels = {
            "kills": "Kills",
            "deaths": "Deaths",
            "kdr": "K/D Ratio",
            "wins": "Wins",
            "losses": "Losses",
            "wlr": "W/L Ratio",
        }
        self.update_buttons()
    
    def update_buttons(self):
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                if child.custom_id == self.current_period:
                    child.style = discord.ButtonStyle.primary
                else:
                    child.style = discord.ButtonStyle.secondary
    
    def get_leaderboard_embed(self, period: str):
        rows = self.periods[period]
        metric_idx = self.metric_indices[self.metric]
        target_row = rows[metric_idx]
        metric_label = self.metric_labels[self.metric]
        
        # Collect all player stats
        leaderboard = []
        for sheet_name in self.wb.sheetnames:
            if sheet_name.casefold() == "sheep wars historical data":
                continue
            try:
                sheet = self.wb[sheet_name]
                value = sheet[f"B{target_row}"].value
                if value is not None and isinstance(value, (int, float)):
                    leaderboard.append((sheet_name, float(value)))
            except Exception:
                continue
        
        # Sort by value descending
        leaderboard.sort(key=lambda x: x[1], reverse=True)
        
        # Build embed
        embed = discord.Embed(
            title=f"{period.title()} {metric_label} Leaderboard",
            color=discord.Color.from_rgb(54, 57, 63)
        )
        
        if not leaderboard:
            embed.description = "No data available"
        else:
            # Top 10
            description_lines = []
            for i, (player, value) in enumerate(leaderboard[:10], 1):
                medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(i, f"**{i}.**")
                description_lines.append(f"{medal} {player}: `{value}`")
            embed.description = "\n".join(description_lines)
        
        return embed
    
    @discord.ui.button(label="Lifetime", custom_id="lifetime", style=discord.ButtonStyle.primary)
    async def lifetime_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_period = "lifetime"
        self.update_buttons()
        embed = self.get_leaderboard_embed(self.current_period)
        await interaction.response.edit_message(embed=embed, view=self)
    
    @discord.ui.button(label="Session", custom_id="session", style=discord.ButtonStyle.secondary)
    async def session_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_period = "session"
        self.update_buttons()
        embed = self.get_leaderboard_embed(self.current_period)
        await interaction.response.edit_message(embed=embed, view=self)
    
    @discord.ui.button(label="Daily", custom_id="daily", style=discord.ButtonStyle.secondary)
    async def daily_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_period = "daily"
        self.update_buttons()
        embed = self.get_leaderboard_embed(self.current_period)
        await interaction.response.edit_message(embed=embed, view=self)
    
    @discord.ui.button(label="Weekly", custom_id="weekly", style=discord.ButtonStyle.secondary)
    async def weekly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_period = "weekly"
        self.update_buttons()
        embed = self.get_leaderboard_embed(self.current_period)
        await interaction.response.edit_message(embed=embed, view=self)
    
    @discord.ui.button(label="Monthly", custom_id="monthly", style=discord.ButtonStyle.secondary)
    async def monthly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_period = "monthly"
        self.update_buttons()
        embed = self.get_leaderboard_embed(self.current_period)
        await interaction.response.edit_message(embed=embed, view=self)


# Create bot with command tree for slash commands
intents = discord.Intents.default()
bot = commands.Bot(command_prefix="!", intents=intents)

# Approval system for verification
class ApprovalView(discord.ui.View):
    def __init__(self, ign: str, requester: str, original_interaction: discord.Interaction):
        super().__init__(timeout=None)
        self.ign = ign
        self.requester = requester
        self.original_interaction = original_interaction
        self.approved = None
        self.done_event = asyncio.Event()
    
    @discord.ui.button(label="Accept", style=discord.ButtonStyle.success)
    async def accept_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.approved = True
        self.done_event.set()
        await interaction.response.edit_message(content=f"You accepted verification for {self.ign}.", view=None)
    
    @discord.ui.button(label="Deny", style=discord.ButtonStyle.danger)
    async def deny_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.approved = False
        self.done_event.set()
        await interaction.response.edit_message(content=f"You denied verification for {self.ign}.", view=None)

# Bot token
# Read from BOT_TOKEN.txt in the same directory
TOKEN_FILE = os.path.join(os.path.dirname(__file__), "BOT_TOKEN.txt")
try:
    with open(TOKEN_FILE, "r", encoding="utf-8") as f:
        DISCORD_TOKEN = f.read().strip()
except Exception as e:
    DISCORD_TOKEN = None
    print(f"[ERROR] Failed to read BOT_TOKEN.txt: {e}")
if not DISCORD_TOKEN:
    raise ValueError("BOT_TOKEN.txt is missing or empty")

@bot.event
async def on_ready():
    print(f"[OK] Bot logged in as {bot.user}")
    try:
        synced = await bot.tree.sync()
        print(f"[OK] Synced {len(synced)} command(s)")
    except Exception as e:
        print(f"[ERROR] Failed to sync commands: {e}")
    # start background scheduler once
    if not getattr(bot, "scheduler_started", False):
        bot.loop.create_task(scheduler_loop())
        bot.scheduler_started = True

@bot.tree.command(name="verify", description="Create a player stats sheet")
@discord.app_commands.describe(ign="Minecraft IGN")
async def verify(interaction: discord.Interaction, ign: str):
    await interaction.response.defer()
    
    try:
        # Get creator user
        creator = None
        if CREATOR_ID is not None:
            try:
                uid = int(CREATOR_ID)
                creator = bot.get_user(uid) or await bot.fetch_user(uid)
            except Exception:
                pass
        
        if creator is None:
            await interaction.followup.send("[ERROR] Cannot reach creator for approval. Contact administrator.")
            return
        
        # Send waiting message to requester
        requester_name = interaction.user.name
        await interaction.followup.send(f"Asked Chuckegg for approval of {ign} verification. Please wait for him to confirm or deny it.")
        
        # Create approval view and send to creator
        view = ApprovalView(ign, requester_name, interaction)
        try:
            await creator.send(f"{requester_name} wants to verify {ign}.", view=view)
        except Exception as e:
            await interaction.followup.send(f"[ERROR] Could not send approval request to creator: {str(e)}")
            return
        
        # Wait for approval (no timeout)
        await view.done_event.wait()
        
        # Process based on approval
        if view.approved:
            result = subprocess.run(
                [sys.executable, "player_stats.py", "-ign", ign],
                capture_output=True,
                text=True,
                timeout=30,
            )

            if result.returncode == 0:
                # add to tracked users list and link Discord account
                added = add_tracked_user(ign)
                link_user_to_ign(interaction.user.id, ign)
                
                # Fetch fresh all-time data (without lifetime flag to update all-time)
                fetch_result = subprocess.run(
                    [sys.executable, "get.py", "-ign", ign],
                    capture_output=True,
                    text=True,
                    timeout=30,
                )
                if fetch_result.returncode != 0:
                    print(f"[WARNING] Failed to fetch fresh data for {ign}: {fetch_result.stderr}")
                
                # Initialize all snapshots (session, daily, weekly, monthly) and deltas in one call
                try:
                    subprocess.run(
                        [sys.executable, "get.py", "-session", "-daily", "-weekly", "-monthly", "-refresh", "-ign", ign],
                        capture_output=True,
                        text=True,
                        timeout=30,
                    )
                    print(f"[OK] Initialized all snapshots for {ign}")
                except Exception as e:
                    print(f"[WARNING] Failed to initialize snapshots for {ign}: {e}")
                
                if added:
                    await interaction.followup.send(f"Chuckegg has accepted the verification of {ign}. {ign} is now verified, linked to your Discord account, and will be automatically tracked daily.")
                else:
                    await interaction.followup.send(f"Chuckegg has accepted the verification of {ign}, but {ign} is already being tracked! Your Discord account has been linked to it.")
            else:
                err = (result.stderr or result.stdout) or "Unknown error"
                await interaction.followup.send(f"Chuckegg has accepted the verification of {ign}, but an error occurred: {sanitize_output(err)}")
        else:
            await interaction.followup.send(f"Chuckegg has denied the verification of {ign}.")
            
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

@bot.tree.command(name="create", description="Create a session snapshot")
@discord.app_commands.describe(ign="Minecraft IGN")
async def create_session(interaction: discord.Interaction, ign: str):
    await interaction.response.defer()
    
    # Check if user is authorized to create session for this username
    if not is_user_authorized(interaction.user.id, ign):
        await interaction.followup.send(f"[ERROR] You are not authorized to create a session for {ign}. Only the user who verified this username can create sessions for it.")
        return
    
    try:
        result = subprocess.run(
            [sys.executable, "create_session.py", "-ign", ign],
            capture_output=True,
            text=True,
            timeout=30,
        )

        if result.returncode == 0:
            await interaction.followup.send(f"Session started for {ign}.")
        else:
            err = (result.stderr or result.stdout) or "Unknown error"
            await interaction.followup.send(f"[ERROR] {sanitize_output(err)}")
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

@bot.tree.command(name="delete", description="Delete your tracked username and all associated data")
@discord.app_commands.describe(ign="Minecraft IGN to delete")
async def delete_user(interaction: discord.Interaction, ign: str):
    await interaction.response.defer()
    
    # Check if user is authorized to delete this username
    if not is_user_authorized(interaction.user.id, ign):
        await interaction.followup.send(f"[ERROR] You are not authorized to delete {ign}. Only the user who verified this username can delete it.")
        return
    
    try:
        EXCEL_FILE = "sheep_wars_stats.xlsx"
        
        # Remove from tracked users
        removed_tracked = remove_tracked_user(ign)
        
        # Remove from user links
        removed_link = unlink_user_from_ign(ign)
        
        # Delete sheet from Excel file
        sheet_deleted = False
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            # Find sheet case-insensitively
            key = ign.casefold()
            found_sheet = None
            for sheet_name in wb.sheetnames:
                if sheet_name.casefold() == key:
                    found_sheet = sheet_name
                    break
            
            if found_sheet:
                del wb[found_sheet]
                wb.save(EXCEL_FILE)
                sheet_deleted = True
        
        if removed_tracked or removed_link or sheet_deleted:
            await interaction.followup.send(f"Successfully deleted all data for {ign}. You are no longer tracked.")
        else:
            await interaction.followup.send(f"[WARNING] No data found for {ign}.")
            
    except Exception as e:
        await interaction.followup.send(f"[ERROR] Failed to delete data: {str(e)}")


@bot.tree.command(name="dmme", description="Send yourself a test DM from the bot")
async def dmme(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    try:
        await interaction.user.send("Hello! This is a private message from the bot.")
        await interaction.followup.send("Sent you a DM.", ephemeral=True)
    except Exception as e:
        await interaction.followup.send("Couldn't DM you. Check your privacy settings (Allow DMs from server members).", ephemeral=True)


@bot.tree.command(name="refresh", description="Manually run daily/weekly/monthly fetch for all tracked users")
@discord.app_commands.describe(mode="One of: daily, weekly, monthly, or all")
@discord.app_commands.choices(mode=[
    discord.app_commands.Choice(name="daily", value="-daily"),
    discord.app_commands.Choice(name="weekly", value="-weekly"),
    discord.app_commands.Choice(name="monthly", value="-monthly"),
    discord.app_commands.Choice(name="all (daily + weekly + monthly)", value="-all"),
])
async def refresh(interaction: discord.Interaction, mode: discord.app_commands.Choice[str]):
    await interaction.response.defer(ephemeral=True)
    try:
        if mode.value == "-all":
            # Single call per user: daily + weekly + monthly + refresh
            flags = ["-daily", "-weekly", "-monthly"]
            fetched = await run_get_for_users_multi(flags)
            if fetched:
                msg = f"Fetched daily, weekly, and monthly for usernames {', '.join(set(fetched))}."
            else:
                msg = "No tracked users to refresh."
        else:
            flag = mode.value
            fetched = await run_get_for_users(flag)
            if fetched:
                msg = f"Fetched {flag} for usernames {', '.join(fetched)}."
            else:
                msg = "No tracked users to refresh."
        
        # Try to DM the invoking user directly
        try:
            await interaction.user.send(msg)
            await interaction.followup.send("Sent you a DM with the results.", ephemeral=True)
        except Exception:
            # Fallback to ephemeral if DMs are closed
            await interaction.followup.send(msg, ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}", ephemeral=True)

@bot.tree.command(name="sheepwars", description="Get player stats with deltas")
@discord.app_commands.describe(ign="Minecraft IGN")
async def sheepwars(interaction: discord.Interaction, ign: str):
    await interaction.response.defer()
    
    try:
        result = subprocess.run(
            [sys.executable, "get.py", "-refresh", "-ign", ign],
            capture_output=True,
            text=True,
            timeout=30
        )
        
        # Read Excel file and get stats
        EXCEL_FILE = "sheep_wars_stats.xlsx"
        if not os.path.exists(EXCEL_FILE):
            await interaction.followup.send("[ERROR] Excel file not found")
            return
        
        wb = load_workbook(EXCEL_FILE)
        
        # Find sheet case-insensitively
        key = ign.casefold()
        found_sheet = None
        for sheet_name in wb.sheetnames:
            if sheet_name.casefold() == key:
                found_sheet = wb[sheet_name]
                break
        
        if found_sheet is None:
            await interaction.followup.send(f"[ERROR] Player sheet '{ign}' not found")
            return
        
        # Pull level and prestige icon for title decoration
        try:
            level_value = int(found_sheet["D40"].value or 0)
        except Exception:
            level_value = 0
        prestige_icon = get_prestige_icon(level_value)

        # Create view with tabs
        view = StatsTabView(found_sheet, ign, level_value, prestige_icon)
        embed = view.get_stats_embed("all-time")
        
        await interaction.followup.send(embed=embed, view=view)
        wb.close()
        
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

@bot.tree.command(name="leaderboard", description="View player leaderboards")
@discord.app_commands.describe(metric="Choose a stat to rank players by")
@discord.app_commands.choices(metric=[
    discord.app_commands.Choice(name="Kills", value="kills"),
    discord.app_commands.Choice(name="Deaths", value="deaths"),
    discord.app_commands.Choice(name="K/D Ratio", value="kdr"),
    discord.app_commands.Choice(name="Wins", value="wins"),
    discord.app_commands.Choice(name="Losses", value="losses"),
    discord.app_commands.Choice(name="W/L Ratio", value="wlr"),
])
async def leaderboard(interaction: discord.Interaction, metric: discord.app_commands.Choice[str]):
    await interaction.response.defer()
    
    try:
        EXCEL_FILE = "sheep_wars_stats.xlsx"
        if not os.path.exists(EXCEL_FILE):
            await interaction.followup.send("[ERROR] Excel file not found")
            return
        
        wb = load_workbook(EXCEL_FILE)
        
        # Create view with period buttons
        view = LeaderboardView(metric.value, wb)
        embed = view.get_leaderboard_embed("lifetime")
        
        await interaction.followup.send(embed=embed, view=view)
        # Note: wb is kept open for the view to use; consider closing after timeout
        
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

# Run bot
if __name__ == "__main__":
    bot.run(DISCORD_TOKEN)
