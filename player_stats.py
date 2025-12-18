import os
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# -------------------
# CLI arguments
# -------------------
parser = argparse.ArgumentParser(description="Create player stats sheet template in Excel")
parser.add_argument("-ign", "--username", required=True, help="Minecraft IGN")
args = parser.parse_args()

USERNAME = args.username
EXCEL_FILE = "sheep_wars_stats.xlsx"

print(f"[INFO] Creating sheet template for {USERNAME}...")

# -------------------
# Create or load workbook
# -------------------
if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
else:
    wb = Workbook()
    # Remove default sheet if present
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

# -------------------
# Create new sheet with player name
# -------------------
sheet_name = USERNAME
# create sheet, but respect existing sheet name casing (case-insensitive match)
existing = None
for s in wb.sheetnames:
    if s.casefold() == sheet_name.casefold():
        existing = s
        break

if existing:
    print(f"[WARNING] Sheet '{existing}' already exists. Removing old data...")
    wb.remove(wb[existing])

ws = wb.create_sheet(USERNAME)

# -------------------
# Styling
# -------------------
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
table_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
table_header_font = Font(bold=True)
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
center_alignment = Alignment(horizontal="center", vertical="center")

# -------------------
# Build tables with headers only
# -------------------
current_row = 1
table_order = ["Session", "Daily", "Weekly", "Monthly", "All-time"]

for table_idx, period in enumerate(table_order):
    # Table title
    ws.merge_cells(f"A{current_row}:F{current_row}")
    title_cell = ws[f"A{current_row}"]
    title_cell.value = f"{period} Stats"
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = center_alignment
    current_row += 1

    # Column headers
    columns = ["Stat", "Value"]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = col_name
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.border = border
        cell.alignment = center_alignment

    current_row += 1

    # Data rows with empty values
    stat_names = ["Kills", "Deaths", "K/D", "Wins", "Losses", "W/L"]

    for stat_name in stat_names:
        ws[f"A{current_row}"] = stat_name
        ws[f"B{current_row}"] = None

        for col in ["A", "B"]:
            cell = ws[f"{col}{current_row}"]
            cell.border = border
            cell.alignment = center_alignment

        current_row += 1

    # Empty row between tables
    current_row += 1

# -------------------
# Set column widths
# -------------------
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 15

# -------------------
# Save workbook
# -------------------
wb.save(EXCEL_FILE)
print(f"[OK] Sheet '{sheet_name}' template created in {EXCEL_FILE}")
