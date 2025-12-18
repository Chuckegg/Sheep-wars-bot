import os
import sys
import argparse
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# -------------------
# CLI arguments
# -------------------
parser = argparse.ArgumentParser(description="Create a session snapshot for a player")
parser.add_argument("-ign", "--username", required=True, help="Minecraft IGN")
parser.add_argument("-firstrun", action="store_true", help="First run - skip clearing existing session data")
args = parser.parse_args()

USERNAME = args.username
FIRST_RUN = args.firstrun
EXCEL_FILE = "sheep_wars_stats.xlsx"

# -------------------
# Fetch stats via get.py
# -------------------
print(f"[LOADING] Fetching stats for {USERNAME} via get.py...")
result = subprocess.run([sys.executable, "get.py", "-ign", USERNAME], 
                       capture_output=True, text=True)

if result.returncode != 0:
    print(f"[ERROR] Error fetching stats:")
    print(result.stderr)
    raise RuntimeError("Failed to fetch stats from get.py")

print(result.stdout)

# -------------------
# Load workbook and store snapshot
# -------------------
if not os.path.exists(EXCEL_FILE):
    print(f"[ERROR] Excel file '{EXCEL_FILE}' not found.")
    raise RuntimeError("Excel file not found")

wb = load_workbook(EXCEL_FILE)

if USERNAME not in wb.sheetnames:
    # try case-insensitive match
    key = USERNAME.casefold()
    found = None
    for s in wb.sheetnames:
        if s.casefold() == key:
            found = s
            break
    if found:
        player_ws = wb[found]
    else:
        print(f"[ERROR] Sheet '{USERNAME}' not found.")
        raise RuntimeError("Player sheet not found")
else:
    player_ws = wb[USERNAME]

# -------------------
# Styling (match other tables)
# -------------------
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
table_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
table_header_font = Font(bold=True)
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
center_alignment = Alignment(horizontal="center", vertical="center")

# Get the latest all-time stats from the All-time Stats section (rows 37-42)
# which were just updated by get.py
all_time_data_start_row = 39  # Data starts after title (37) and headers (38)

snapshot_data = [
    (player_ws[f"B{all_time_data_start_row}"].value, "Kills"),
    (player_ws[f"B{all_time_data_start_row + 1}"].value, "Deaths"),
    (player_ws[f"B{all_time_data_start_row + 2}"].value, "K/D"),
    (player_ws[f"B{all_time_data_start_row + 3}"].value, "Wins"),
    (player_ws[f"B{all_time_data_start_row + 4}"].value, "Losses"),
    (player_ws[f"B{all_time_data_start_row + 5}"].value, "W/L"),
]

# Check if Session Start table already exists by checking for merged cells
session_start_exists = False
for merged_range in player_ws.merged_cells.ranges:
    if "D1" in merged_range:
        session_start_exists = True
        break

if not session_start_exists:
    # Create the table structure
    # Add title in row 1 (merged across D-E)
    player_ws.merge_cells("D1:E1")
    title_cell = player_ws["D1"]
    title_cell.value = "Session Start"
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = center_alignment

# Add column headers in row 2 (ensure they exist and match get.py formatting)
cols = ["Snapshot", "Value"]
for col_idx, col_name in enumerate(cols):
    col_letter = chr(68 + col_idx)  # D=68, E=69 in ASCII
    cell = player_ws[f"{col_letter}2"]
    cell.value = col_name
    cell.font = table_header_font
    cell.fill = table_header_fill
    cell.border = border
    cell.alignment = center_alignment

# Clear previous session stats (reset) - skip on first run
if not FIRST_RUN:
    for row in range(3, 9):
        player_ws[f"B{row}"] = None

# Update snapshot values in rows 3-8
for idx, (stat_value, stat_name) in enumerate(snapshot_data):
    row = 3 + idx
    player_ws[f"D{row}"] = stat_name
    player_ws[f"E{row}"] = stat_value
    
    # Apply formatting to snapshot cells
    for col in ["D", "E"]:
        cell = player_ws[f"{col}{row}"]
        cell.border = border
        cell.alignment = center_alignment

wb.save(EXCEL_FILE)
print(f"[OK] Session snapshot created for {USERNAME}")

