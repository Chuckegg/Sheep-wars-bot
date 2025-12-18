import requests
from bs4 import BeautifulSoup
from datetime import datetime
import re
import os
import argparse
import time
import random
import json
import gzip
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Get script directory for file operations
SCRIPT_DIR = Path(__file__).parent.absolute()

# -------------------
# CLI arguments
# -------------------
parser = argparse.ArgumentParser(description="Fetch Sheep Wars stats")
parser.add_argument("-ign", "--username", required=True, help="Minecraft IGN")
parser.add_argument("-nolifetime", action="store_true", help="Don't update all-time stats in player sheet")
parser.add_argument("-session", action="store_true", help="Log snapshot into Session Start section")
parser.add_argument("-daily", action="store_true", help="Log snapshot into Daily Stats section")
parser.add_argument("-weekly", action="store_true", help="Log snapshot into Weekly Stats section")
parser.add_argument("-monthly", action="store_true", help="Log snapshot into Monthly Stats section")
parser.add_argument("-refresh", action="store_true", help="Refresh all stats with deltas from snapshots")
parser.add_argument("-proxy", action="store_true", help="Use proxy rotation from ProxyScrape")
parser.add_argument("-noproxy", action="store_true", help="Disable proxies (direct connection only)")
args = parser.parse_args()

USERNAME = args.username
URL = f"https://plancke.io/hypixel/player/stats/{USERNAME}"

# Rotating user agents to appear more like different browsers
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) Gecko/20100101 Firefox/133.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.2 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
]

HEADERS = {
    "User-Agent": random.choice(USER_AGENTS),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "DNT": "1",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Cache-Control": "max-age=0",
}

EXCEL_FILE = str(SCRIPT_DIR / "sheep_wars_stats.xlsx")
SHEET_NAME = "Sheep Wars historical data"
PROXY_CACHE_FILE = str(SCRIPT_DIR / "proxy_cache.json")
PROXYSCRAPE_API_KEY = os.environ.get("PROXYSCRAPE_API_KEY") or os.environ.get("PROXYSCRAPE_KEY") or "f3g7edlwjly872gzdaai"

# -------------------
# Proxy Management
# -------------------
def fetch_proxies_from_proxyscrape():
    """Fetch proxies, preferring ProxyScrape with API key; fallback to free sources"""
    try:
        print("[PROXY] Fetching proxies...")

        candidates = []
        # Paid/API-keyed endpoints first (if key provided). Avoid premium-only 'format' param.
        if PROXYSCRAPE_API_KEY:
            candidates.append(
                f"https://api.proxyscrape.com/v2/?request=displayproxies&protocol=http&timeout=15000&country=all&ssl=all&anonymity=all&apikey={PROXYSCRAPE_API_KEY}"
            )
            candidates.append(
                f"https://api.proxyscrape.com/v2/?request=displayproxies&protocol=https&timeout=15000&country=all&ssl=all&anonymity=all&apikey={PROXYSCRAPE_API_KEY}"
            )

        # Free endpoint fallback (no apikey, no 'format' param)
        candidates.append("https://api.proxyscrape.com/v2/?request=displayproxies&protocol=http&timeout=10000&country=all&ssl=all&anonymity=all")
        candidates.append("https://api.proxyscrape.com/v2/?request=displayproxies&protocol=https&timeout=10000&country=all&ssl=all&anonymity=all")
        # Secondary free source (plaintext list)
        candidates.append("https://www.proxy-list.download/api/v1/get?type=http")

        for url in candidates:
            try:
                response = requests.get(url, timeout=15)
                if response.status_code != 200:
                    print(f"[PROXY] {url} -> HTTP {response.status_code}")
                    continue

                text = response.text.strip()

                # Parse plaintext for both ProxyScrape and proxy-list.download
                if 'proxy-list.download' in url:
                    # Endpoint returns plaintext host:port per line
                    proxies = [p.strip() for p in text.split('\n') if p.strip() and ':' in p]
                else:
                    # ProxyScrape endpoints default to plaintext if 'format' is omitted
                    # Filter out any error messages
                    if 'format are premium features' in text.lower():
                        proxies = []
                    else:
                        proxies = [p.strip() for p in text.split('\n') if p.strip() and ':' in p and not p.startswith('{')]

                if proxies:
                    print(f"[PROXY] Fetched {len(proxies)} proxies from {url.split('/')[2]}")
                    return proxies
                else:
                    print(f"[PROXY] No proxies returned from {url.split('/')[2]} (first 120 chars: {text[:120]!r})")
            except Exception as e:
                print(f"[PROXY] Error calling {url}: {e}")
                continue

        print("[PROXY] Could not fetch proxies from any source")
        return []

    except Exception as e:
        print(f"[PROXY] Error fetching proxies: {e}")
        return []

def load_cached_proxies():
    """Load proxies from cache file"""
    if os.path.exists(PROXY_CACHE_FILE):
        try:
            with open(PROXY_CACHE_FILE, 'r') as f:
                data = json.load(f)
                # Check if cache is less than 1 hour old
                if time.time() - data.get('timestamp', 0) < 3600:
                    print(f"[PROXY] Loaded {len(data.get('proxies', []))} cached proxies")
                    return data.get('proxies', [])
        except Exception as e:
            print(f"[PROXY] Error loading cache: {e}")
    return []

def save_proxy_cache(proxies):
    """Save working proxies to cache"""
    try:
        with open(PROXY_CACHE_FILE, 'w') as f:
            json.dump({
                'timestamp': time.time(),
                'proxies': proxies
            }, f)
    except Exception as e:
        print(f"[PROXY] Error saving cache: {e}")

def test_proxy(proxy, test_url="https://httpbin.org/ip", timeout=10):
    """Test if a proxy supports HTTPS CONNECT by performing an HTTPS request."""
    try:
        proxy_dict = {
            'http': f'http://{proxy}',
            'https': f'http://{proxy}',
        }
        r = requests.get(test_url, proxies=proxy_dict, timeout=timeout)
        return r.status_code == 200
    except Exception:
        return False

def get_working_proxies(max_test=30):
    """Get a list of working HTTPS-capable proxies.
    Tries cache first; if none work, fetches fresh list and retries.
    """
    def test_list(proxies_list):
        if not proxies_list:
            return []
        print(f"[PROXY] Testing up to {max_test} proxies for HTTPS...")
        random.shuffle(proxies_list)
        found = []
        for proxy in proxies_list[:max_test]:
            if test_proxy(proxy):
                found.append(proxy)
                print(f"[PROXY] [OK] {proxy}")
                if len(found) >= 5:
                    break
            else:
                print(f"[PROXY] [FAIL] {proxy}")
        return found

    # Try cache
    proxies = load_cached_proxies()
    used_cache = bool(proxies)
    working = test_list(proxies)

    # If cache failed, fetch fresh list and retry
    if not working:
        if used_cache:
            print("[PROXY] Cached proxies failed; fetching fresh list...")
        proxies = fetch_proxies_from_proxyscrape()
        working = test_list(proxies)

    if working:
        save_proxy_cache(working)
    return working

# Initialize proxy pool if enabled
PROXY_POOL = []
if args.proxy and not args.noproxy:
    PROXY_POOL = get_working_proxies()
    if PROXY_POOL:
        print(f"[PROXY] Ready with {len(PROXY_POOL)} working proxies")
    else:
        print("[PROXY] No working proxies found, will use direct connection")

# -------------------
# Fetch page with retry logic
# -------------------
def fetch_with_retry(url, headers, max_retries=3, initial_delay=2, use_proxies=True, request_timeout=20):
    """Fetch URL with exponential backoff retry logic and optional proxy rotation"""
    # Create a session for better request handling
    session = requests.Session()
    proxies_to_try = PROXY_POOL.copy() if (use_proxies and PROXY_POOL) else [None]
    
    for attempt in range(max_retries):
        # Rotate through proxies
        proxy = None
        proxy_dict = None
        
        if proxies_to_try and proxies_to_try[0] is not None:
            proxy = random.choice(proxies_to_try)
            proxy_dict = {
                'http': f'http://{proxy}',
                'https': f'http://{proxy}'
            }
            print(f"  Using proxy: {proxy}")
        
        try:
            # Random delay between requests
            if attempt > 0:
                delay = initial_delay * (2 ** attempt) + random.uniform(0, 1)
                print(f"  Retry {attempt}/{max_retries} - waiting {delay:.1f}s...")
                time.sleep(delay)
            else:
                # Small random delay even on first attempt to appear more human
                time.sleep(random.uniform(0.5, 2.0))
            
            response = session.get(url, headers=headers, proxies=proxy_dict, timeout=request_timeout)
            response.raise_for_status()
            return response
            
        except requests.exceptions.RequestException as e:
            print(f"  Request failed: {e}")
            
            # If proxy failed, remove it from the list and try another
            if proxy and proxy in proxies_to_try:
                proxies_to_try.remove(proxy)
                print(f"  Removing failed proxy: {proxy}")
            
            # If no more proxies or on last attempt, try direct connection
            if not proxies_to_try or attempt == max_retries - 1:
                if proxy_dict:  # We were using proxies, try direct now
                    print("  Trying direct connection...")
                    proxies_to_try = [None]
                    continue
                else:
                    raise
        finally:
            session.close()
    
    return None

response = fetch_with_retry(URL, HEADERS)
if response is None:
    raise RuntimeError("Network fetch failed after retries (proxies + direct). Try again later or use -noproxy.")
# requests handles gzip automatically with response.text
soup = BeautifulSoup(response.text, "html.parser")

text = soup.get_text("\n")

# -------------------
# Extract stats
# -------------------
pattern = re.compile(
    r"Sheep Wars.*?"
    r"Wins:\s*([\d,]+).*?"
    r"Losses:\s*([\d,]+).*?"
    r"W/L:\s*([\d.]+).*?"
    r"Kills:\s*([\d,]+).*?"
    r"Deaths:\s*([\d,]+).*?"
    r"K/D:\s*([\d.]+)",
    re.S
)

match = pattern.search(text)

if not match:
    print(f"[ERROR] Sheep Wars stats NOT found for {USERNAME}")
    print(f"[DEBUG] Page length: {len(text)} characters")
    print(f"[DEBUG] First 500 chars of page:")
    print(text[:500])
    idx = text.find("Sheep Wars")
    if idx != -1:
        print(f"\n[DEBUG] Found 'Sheep Wars' at position {idx}")
        print(text[idx:idx + 800])
    else:
        print(f"\n[DEBUG] 'Sheep Wars' text not found in page")
    raise RuntimeError("Extraction failed")

wins, losses, wl, kills, deaths, kd = match.groups()

# Extract Wool and Level (separate from Sheep Wars stats)
wool_pattern = re.compile(r"Wool:\s*([\d,]+)", re.S)
# Prefer the Level that appears immediately after the Wool stat (the Sheep Wars level),
# fall back to the first Level match if not found.
level_pattern = re.compile(r"Level:\s*([\d,]+)", re.S)

wool_match = wool_pattern.search(text)
level_match = None

if wool_match:
    # Look for a Level entry after the Wool line (within a reasonable window)
    level_after_wool = level_pattern.search(text, wool_match.end())
    if level_after_wool:
        level_match = level_after_wool

# Fallback to the first Level match anywhere in the page
if level_match is None:
    level_match = level_pattern.search(text)

wool = wool_match.group(1) if wool_match else "0"
level = level_match.group(1) if level_match else "0"

# -------------------
# Terminal output
# -------------------
if not args.nolifetime:
    print(f"[OK] Sheep Wars stats extracted for {USERNAME}:")
    print(f"  Wins   : {wins}")
    print(f"  Losses : {losses}")
    print(f"  W/L    : {wl}")
    print(f"  Kills  : {kills}")
    print(f"  Deaths : {deaths}")
    print(f"  K/D    : {kd}")
    print(f"  Wool   : {wool}")
    print(f"  Level  : {level}")

# -------------------
# Prepare Excel row
# -------------------
headers = ["Date/Time", "Username", "Kills", "Deaths", "K/D", "Wins", "Losses", "W/L"]

row = [
    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    USERNAME,
    int(kills.replace(",", "")),
    int(deaths.replace(",", "")),
    float(kd),
    int(wins.replace(",", "")),
    int(losses.replace(",", "")),
    float(wl),
]

# -------------------
# Create or load workbook
# -------------------
if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
else:
    wb = Workbook()

# -------------------
# Create or select sheet
# -------------------
if SHEET_NAME in wb.sheetnames:
    ws = wb[SHEET_NAME]
else:
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(headers)  # write headers once

# -------------------
# Append row
# -------------------
ws.append(row)

def find_sheet_case_insensitive(workbook, name):
    key = name.casefold()
    for s in workbook.sheetnames:
        if s.casefold() == key:
            return s
    return None


# -------------------
# Update player's personal sheet (All-time stats)
# -------------------
found = find_sheet_case_insensitive(wb, USERNAME)
if found:
    player_ws = wb[found]
    
    # Update All-time stats if not disabled
    if not args.nolifetime:
        # Find the "All-time Stats" table and update its values
        # Tables start at row 1 with structure: Title, Headers, then 6 stat rows
        # Session (rows 1-9), Daily (rows 11-19), Weekly (rows 21-29), Monthly (rows 31-39), All-time (rows 37-45)
        
        all_time_start_row = 37  # All-time Stats section starts at row 37
        all_time_data_start_row = all_time_start_row + 2  # Data rows start after title and headers
        
        # Map stat names to row offsets and values
        stat_mapping = {
            "Kills": (all_time_data_start_row, int(kills.replace(",", ""))),
            "Deaths": (all_time_data_start_row + 1, int(deaths.replace(",", ""))),
            "K/D": (all_time_data_start_row + 2, float(kd)),
            "Wins": (all_time_data_start_row + 3, int(wins.replace(",", ""))),
            "Losses": (all_time_data_start_row + 4, int(losses.replace(",", ""))),
            "W/L": (all_time_data_start_row + 5, float(wl)),
        }
        
        # Update the values in column B
        for stat_name, (row_num, value) in stat_mapping.items():
            player_ws[f"B{row_num}"] = value
        
        # Update Wool and Level in D39 and D40
        player_ws["D39"] = int(wool.replace(",", ""))
        player_ws["D40"] = int(level.replace(",", ""))
        
        if not args.nolifetime:
            print(f"[OK] All-time stats updated in sheet '{USERNAME}'")
            print(f"[OK] Wool: {wool}, Level: {level} saved to D39:D40")
    else:
        if not args.nolifetime:
            print(f"[SKIP] Skipped all-time stats update (-nolifetime flag)")
    
    # Update Session stats (calculate difference from snapshot)
    session_data_start_row = 3  # Session stats data rows start at row 3
    
    # Check if snapshot exists
    if player_ws["E3"].value is not None:
        snapshot_values = {
            "Kills": player_ws["E3"].value,
            "Deaths": player_ws["E4"].value,
            "K/D": player_ws["E5"].value,
            "Wins": player_ws["E6"].value,
            "Losses": player_ws["E7"].value,
            "W/L": player_ws["E8"].value,
        }
        
        # Calculate session stats as deltas (current - snapshot)
        current_kills = int(kills.replace(",", ""))
        current_deaths = int(deaths.replace(",", ""))
        current_wins = int(wins.replace(",", ""))
        current_losses = int(losses.replace(",", ""))

        session_kills = (current_kills - (snapshot_values["Kills"] or 0)) or 0
        session_deaths = (current_deaths - (snapshot_values["Deaths"] or 0)) or 0
        session_wins = (current_wins - (snapshot_values["Wins"] or 0)) or 0
        session_losses = (current_losses - (snapshot_values["Losses"] or 0)) or 0

        # Compute ratios from the deltas (not delta of ratios)
        if session_deaths and session_deaths != 0:
            session_kd = round(session_kills / session_deaths, 2)
        else:
            session_kd = float(session_kills) if session_kills else 0.0

        if session_losses and session_losses != 0:
            session_wl = round(session_wins / session_losses, 2)
        else:
            session_wl = float(session_wins) if session_wins else 0.0
        
        # Update session stats in column B (rows 3-8)
        player_ws["B3"] = session_kills
        player_ws["B4"] = session_deaths
        player_ws["B5"] = session_kd
        player_ws["B6"] = session_wins
        player_ws["B7"] = session_losses
        player_ws["B8"] = session_wl
        
        if not args.nolifetime:
            print(f"[OK] Session stats updated for '{USERNAME}'")
    else:
        # No snapshot found - create it automatically using current all-time stats
        if not args.nolifetime:
            print(f"[INFO] No session snapshot found. Creating one now...")
            # Write snapshot to D3:E8 (row 1 is title, row 2 is headers)
            snapshot_vals = [
                int(kills.replace(",", "")),
                int(deaths.replace(",", "")),
                float(kd),
                int(wins.replace(",", "")),
                int(losses.replace(",", "")),
                float(wl),
            ]
            stat_names = ["Kills", "Deaths", "K/D", "Wins", "Losses", "W/L"]
            for idx, stat_name in enumerate(stat_names):
                r = 3 + idx  # Data starts at row 3
                player_ws[f"D{r}"] = stat_name
                player_ws[f"E{r}"] = snapshot_vals[idx]
            print(f"[OK] Session snapshot created for '{USERNAME}'")

    # -------------------
    # Helper: ensure snapshot table exists and write values into D/E for a given section
    # -------------------
    def write_section_snapshot(title_row, header_row, data_start_row, title_text):
        # Only write snapshot data rows into D/E to avoid touching merged title/header cells.
        # ensure headers exist (Snapshot / Value) and are styled
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font_local = Font(bold=True)
        # write column headers (do not touch merged title cells)
        player_ws[f"D{header_row}"].value = "Snapshot"
        player_ws[f"E{header_row}"].value = "Value"
        player_ws[f"D{header_row}"].font = header_font_local
        player_ws[f"E{header_row}"].font = header_font_local
        player_ws[f"D{header_row}"].fill = header_fill
        player_ws[f"E{header_row}"].fill = header_fill
        player_ws[f"D{header_row}"].alignment = Alignment(horizontal="center", vertical="center")
        player_ws[f"E{header_row}"].alignment = Alignment(horizontal="center", vertical="center")
        player_ws[f"D{header_row}"].border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        player_ws[f"E{header_row}"].border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        vals = [
            int(kills.replace(",", "")),
            int(deaths.replace(",", "")),
            float(kd),
            int(wins.replace(",", "")),
            int(losses.replace(",", "")),
            float(wl),
        ]

        stat_names = ["Kills", "Deaths", "K/D", "Wins", "Losses", "W/L"]
        for idx, stat_name in enumerate(stat_names):
            r = data_start_row + idx
            player_ws[f"D{r}"] = stat_name
            player_ws[f"E{r}"] = vals[idx]

            # apply simple formatting
            cell_d = player_ws[f"D{r}"]
            cell_e = player_ws[f"E{r}"]
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
            cell_e.alignment = Alignment(horizontal="center", vertical="center")
            cell_d.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            cell_e.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Write snapshots to requested sections
    if args.session:
        # Session Start at row 1, headers 2, data 3-8
        write_section_snapshot(1, 2, 3, "Session Start")
    if args.daily:
        # Daily title at row 10, headers 11, data 12-17
        write_section_snapshot(10, 11, 12, "Daily Start")
    if args.weekly:
        # Weekly title at row 19, headers 20, data 21-26
        write_section_snapshot(19, 20, 21, "Weekly Start")
    if args.monthly:
        # Monthly title at row 28, headers 29, data 30-35
        write_section_snapshot(28, 29, 30, "Monthly Start")

else:
    if not args.nolifetime:
        print(f"[WARNING] Sheet '{USERNAME}' not found. Create it first with player_stats.py")

# -------------------
# Refresh: Compute deltas and write into column B (if -refresh flag set)
# -------------------
if args.refresh and USERNAME in wb.sheetnames:
    player_ws = wb[USERNAME]
    
    # helper to read snapshot values from D/E
    def read_snapshot(start_row):
        stat_names = ["Kills", "Deaths", "K/D", "Wins", "Losses", "W/L"]
        out = {}
        for i, name in enumerate(stat_names):
            val = player_ws[f"E{start_row + i}"].value
            if val is None:
                return None
            out[name] = val
        return out
    
    # read current all-time values from B (rows 39-44)
    stat_names = ["Kills", "Deaths", "K/D", "Wins", "Losses", "W/L"]
    all_time_start_row = 39
    all_time = {}
    for i, name in enumerate(stat_names):
        all_time[name] = player_ws[f"B{all_time_start_row + i}"].value or 0
    
    # periods: (snapshot_start_row, target_start_row)
    periods = {
        "Session": (3, 3),
        "Daily": (12, 12),
        "Weekly": (21, 21),
        "Monthly": (30, 30),
    }
    
    for period, (snap_row, target_row) in periods.items():
        snap = read_snapshot(snap_row)
        if snap is None:
            continue
        
        # compute deltas
        try:
            kills_delta = (all_time.get("Kills", 0) or 0) - (snap.get("Kills", 0) or 0)
            deaths_delta = (all_time.get("Deaths", 0) or 0) - (snap.get("Deaths", 0) or 0)
            wins_delta = (all_time.get("Wins", 0) or 0) - (snap.get("Wins", 0) or 0)
            losses_delta = (all_time.get("Losses", 0) or 0) - (snap.get("Losses", 0) or 0)
        except Exception:
            kills_delta = deaths_delta = wins_delta = losses_delta = 0
        
        # write counts and ratios
        player_ws[f"B{target_row + 0}"] = kills_delta
        player_ws[f"B{target_row + 1}"] = deaths_delta
        
        # K/D ratio
        if deaths_delta and deaths_delta != 0:
            kd_ratio = kills_delta / deaths_delta
        else:
            kd_ratio = float(kills_delta) if kills_delta else 0.0
        player_ws[f"B{target_row + 2}"] = round(kd_ratio, 2)
        
        # W/L ratio
        player_ws[f"B{target_row + 3}"] = wins_delta
        player_ws[f"B{target_row + 4}"] = losses_delta
        if losses_delta and losses_delta != 0:
            wl_ratio = wins_delta / losses_delta
        else:
            wl_ratio = float(wins_delta) if wins_delta else 0.0
        player_ws[f"B{target_row + 5}"] = round(wl_ratio, 2)
    
    # Write all-time current values
    for i, name in enumerate(stat_names):
        player_ws[f"B{all_time_start_row + i}"] = all_time.get(name, 0)

# -------------------
# Save workbook
# -------------------
wb.save(EXCEL_FILE)

if not args.nolifetime:
    print(f"[DATA] Data written to {EXCEL_FILE}")
